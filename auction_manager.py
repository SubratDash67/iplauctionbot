# auction_manager.py
"""
Auction Manager Module - v2.1
Handles all auction-related logic with atomic operations, auto-bid, and anti-sniping
"""

import asyncio
import time
import logging
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field

# Get logger from main bot module
logger = logging.getLogger("AuctionBot.Manager")

from config import (
    DEFAULT_BASE_PRICE,
    DEFAULT_COUNTDOWN,
    COUNTDOWN_GAP,
    PLAYER_GAP,
    DEFAULT_AUCTION_FILE,
    get_bid_increment,
)
from database import Database
from utils import FileManager, MessageFormatter, format_amount
from retained_players import RETAINED_PLAYERS, get_remaining_purse


@dataclass
class BidResult:
    """Result of a bid attempt"""

    success: bool
    message: str
    amount: int = 0
    team: str = ""
    is_auto_bid: bool = False
    auto_bids_triggered: List[dict] = field(default_factory=list)
    original_bid_amount: int = 0


class AuctionManager:
    """
    Main auction management class with atomic operations.
    """

    def __init__(
        self, teams: Dict[str, int], excel_file: str, db_path: str = "auction.db"
    ):
        self.db = Database(db_path)
        self.excel_file = excel_file
        self.file_manager = FileManager()
        self.formatter = MessageFormatter()

        # Configuration settings
        self.countdown_gap = COUNTDOWN_GAP
        self.player_gap = PLAYER_GAP

        # Initialize teams with retained players deducted - ONLY if no existing teams
        adjusted_teams = {}
        for team_code, initial_purse in teams.items():
            remaining = get_remaining_purse(team_code, initial_purse)
            adjusted_teams[team_code] = remaining

        # Only initialize teams if database is empty (preserves data across bot restarts)
        self.db.init_teams_if_empty(adjusted_teams)

        self._bid_lock = asyncio.Lock()

        # In-memory state (synced with DB)
        self._load_state_from_db()

        # NOTE: Automatic loading of retained players and Excel generation has been removed
        # per request. Use the /loadretained command to initialize this data.

    def load_retained_data(self) -> Tuple[bool, str]:
        """
        Loads retained players into the database and initializes the Excel file.
        This serves as the /loadretained command handler.
        Safe to call multiple times - skips players that already exist.
        """
        try:
            # 1. Populate DB with retained players (handles duplicates internally)
            count = self._initialize_retained_players()

            # 2. Get fresh data from DB
            teams = self.db.get_teams()
            squads = self.db.get_all_squads()

            # 3. Initialize/Update Excel (regenerate from DB to ensure consistency)
            self.file_manager.initialize_excel_with_retained_players(
                self.excel_file, teams, squads
            )

            if count > 0:
                return True, f"Added {count} retained players. Excel updated."
            else:
                return True, "All retained players already loaded. Excel refreshed."
        except Exception as e:
            logger.error(f"Error loading retained data: {e}")
            return False, f"Error loading retained data: {str(e)}"

    def _initialize_retained_players(self) -> int:
        """Initialize retained players into squads.
        Checks globally across ALL teams to prevent duplicates.
        Updates overseas status for existing players.
        """
        existing_squads = self.db.get_all_squads()

        # Build a global set of all existing players (case-insensitive)
        all_existing_players = set()
        for team_code, squad in existing_squads.items():
            for item in squad:
                all_existing_players.add(item[0].lower())

        count = 0
        for team_code, players in RETAINED_PLAYERS.items():
            for entry in players:
                # FIX: Handle both (Name, Price) and (Name, Price, IsOverseas) formats
                if len(entry) == 3:
                    player_name, salary, is_overseas = entry
                else:
                    player_name, salary = entry
                    is_overseas = False

                # Force update status in DB even if player exists
                self.db.update_overseas_status(player_name, is_overseas)

                if player_name.lower() not in all_existing_players:
                    try:
                        success = self.db.add_to_squad(
                            team_code,
                            player_name,
                            salary,
                            acquisition_type="retained",
                            is_overseas=is_overseas,
                        )
                        if success:
                            all_existing_players.add(player_name.lower())
                            count += 1
                    except Exception as e:
                        logger.warning(
                            f"Could not add retained player {player_name}: {e}"
                        )
        return count

    def _load_auction_excel(
        self, filepath: str, num_sets: Optional[int] = None
    ) -> Tuple[bool, str]:
        """Load NEXT N sets from Excel file (Auction_list.xlsx).

        Args:
            filepath: Path to the Excel file
            num_sets: Number of NEW sets to load (loads sets from current_max+1 to current_max+num_sets)
        """
        from openpyxl import load_workbook

        try:
            # Get currently loaded max set number
            current_max_loaded = self.db.get_max_loaded_set()

            # Calculate the range of sets to load
            start_set = current_max_loaded + 1
            end_set = current_max_loaded + num_sets if num_sets else 67

            # Get existing lists to check for special lists (accelerated, released)
            existing_lists = self.db.get_player_lists()
            existing_set_names = set(existing_lists.keys()) if existing_lists else set()

            # Special lists that should not be treated as regular sets
            special_lists = {
                "accelerated",
                "released",
                "released players",
                "unsold players",
                "skipped",
            }

            players_by_set = {}
            set_number_map = {}  # Stores {set_name: set_number} for sorting
            row_count = 0
            skipped_count = 0
            max_set_loaded = current_max_loaded

            # Load Excel workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active

            # Find header row and column indices
            header_row = None
            col_indices = {}

            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1
            ):
                if row is None:
                    continue
                # Check if this row contains headers
                row_str = [str(cell).strip().upper() if cell else "" for cell in row]

                # Look for key columns - more flexible matching
                for col_idx, cell_val in enumerate(row_str):
                    # Match "PLAYER NO" or "PLAYER NO."
                    if "PLAYER NO" in cell_val or "PLAYER_NO" in cell_val:
                        col_indices["player_no"] = col_idx
                    # Match "SET NO" or "SET NO." - check this BEFORE "SET"
                    elif "SET NO" in cell_val or "SET_NO" in cell_val:
                        col_indices["set_no"] = col_idx
                    # Match "SET" but not "SET NO"
                    elif cell_val == "SET":
                        col_indices["set"] = col_idx
                    # Match "PLAYER" exactly (not "PLAYER NO")
                    elif cell_val == "PLAYER":
                        col_indices["player"] = col_idx
                    # Match "BASE" or "BASE PRICE"
                    elif "BASE" in cell_val:
                        col_indices["base"] = col_idx

                # If we found the key columns, this is the header row
                if "set_no" in col_indices and "player" in col_indices:
                    header_row = row_idx
                    logger.info(f"Found header row at {row_idx}: {col_indices}")
                    break

            if header_row is None:
                # Default to first row as header, standard column positions
                # Based on user's columns: A=Player NO., B=SET NO, C=SET, D=PLAYER, E=BASE
                header_row = 1
                col_indices = {
                    "player_no": 0,  # Column A
                    "set_no": 1,  # Column B
                    "set": 2,  # Column C
                    "player": 3,  # Column D
                    "base": 4,  # Column E
                }
                logger.info(f"Using default column mapping: {col_indices}")

            # Process data rows
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                row_count += 1

                if row is None or not any(row):
                    skipped_count += 1
                    continue

                # Extract values using column indices
                set_no_val = (
                    row[col_indices.get("set_no", 1)]
                    if col_indices.get("set_no", 1) < len(row)
                    else None
                )
                set_name_val = (
                    row[col_indices.get("set", 2)]
                    if col_indices.get("set", 2) < len(row)
                    else None
                )
                player_name_val = (
                    row[col_indices.get("player", 3)]
                    if col_indices.get("player", 3) < len(row)
                    else None
                )
                base_val = (
                    row[col_indices.get("base", 4)]
                    if col_indices.get("base", 4) < len(row)
                    else None
                )

                # Validate player name
                if not player_name_val:
                    skipped_count += 1
                    continue

                player_name_raw = str(player_name_val).strip()
                if not player_name_raw or player_name_raw.upper() in (
                    "PLAYER",
                    "NAME",
                    "PLAYER NAME",
                ):
                    skipped_count += 1
                    continue

                # Check if player is overseas (has ✈️ emoji)
                is_overseas = "✈️" in player_name_raw
                # Remove emoji from name for storage
                player_name = player_name_raw.replace("✈️", "").strip()

                # Parse Set Number
                if set_no_val is None:
                    skipped_count += 1
                    continue

                try:
                    set_number = int(float(str(set_no_val).replace(",", "")))
                except (ValueError, TypeError):
                    skipped_count += 1
                    continue

                # Apply set range filter: only load sets in range [start_set, end_set]
                if set_number < start_set or set_number > end_set:
                    skipped_count += 1
                    continue

                # Track max set number loaded
                if set_number > max_set_loaded:
                    max_set_loaded = set_number

                # Determine Set Name (use SET column, e.g., "BA1")
                if set_name_val:
                    set_key = str(set_name_val).strip()
                else:
                    set_key = f"Set {set_number}"

                # Skip if this set already exists (as a regular set, not special list)
                if (
                    set_key.lower() in existing_set_names
                    and set_key.lower() not in special_lists
                ):
                    skipped_count += 1
                    continue

                # Parse Base Price (in Lakhs -> convert to Rupees)
                base_price = DEFAULT_BASE_PRICE
                if base_val is not None:
                    try:
                        clean_val = str(base_val).replace(",", "").strip()
                        if clean_val:
                            lakhs = float(clean_val)
                            # Convert Lakhs to Rupees (1 Lakh = 100,000)
                            base_price = int(lakhs * 100000)
                    except (ValueError, TypeError):
                        base_price = DEFAULT_BASE_PRICE

                # Group players by set
                if set_key not in players_by_set:
                    players_by_set[set_key] = []
                    set_number_map[set_key] = set_number
                    try:
                        self.db.create_list(set_key)
                    except Exception:
                        pass

                players_by_set[set_key].append((player_name, base_price, is_overseas))

            wb.close()

            if not players_by_set:
                if current_max_loaded >= 67:
                    return (
                        True,
                        f"All 67 sets have been loaded. Use `/clear` to reload from beginning.",
                    )
                # Provide more diagnostic info
                debug_info = f"Header row: {header_row}, Columns: {col_indices}, Rows scanned: {row_count}, Skipped: {skipped_count}"
                logger.warning(f"No sets loaded. {debug_info}")
                return (
                    False,
                    f"No new sets found in range {start_set}-{end_set}. Currently at set {current_max_loaded}.\n"
                    f"Debug: {debug_info}",
                )

            total_players = 0
            total_skipped = 0
            for set_name, players in players_by_set.items():
                # Players now have format: (name, base_price, is_overseas)
                added, skipped = self.db.add_players_to_list_with_overseas(
                    set_name, players
                )
                total_players += added
                total_skipped += skipped

            # Update list order to include new sets (merge with existing)
            # Keep special lists at their positions
            if players_by_set:
                existing_order = self.db.get_list_order()
                existing_order_lower = {s.lower() for s in existing_order}

                # Separate special lists from regular sets
                special_at_start = []  # released goes at start
                special_at_end = []  # accelerated, skipped go at end
                regular_sets = []

                for list_name in existing_order:
                    if list_name.lower() == "released":
                        special_at_start.append(list_name)
                    elif list_name.lower() in special_lists:
                        special_at_end.append(list_name)
                    else:
                        regular_sets.append(list_name)

                # Sort new sets by their set number
                new_sets_sorted = sorted(
                    players_by_set.keys(), key=lambda k: set_number_map.get(k, 999)
                )

                # Add new sets to regular sets list (avoid duplicates)
                for s in new_sets_sorted:
                    if s.lower() not in existing_order_lower:
                        regular_sets.append(s)

                # Final order: released (if exists) -> regular sets -> accelerated/skipped
                final_order = special_at_start + regular_sets + special_at_end
                self.db.set_list_order(final_order)

            # Update max loaded set in database
            if max_set_loaded > current_max_loaded:
                self.db.set_max_loaded_set(max_set_loaded)

            msg = f"Loaded {total_players} players from {len(players_by_set)} NEW sets (sets {start_set}-{max_set_loaded})"
            if total_skipped > 0:
                msg += f"\n⚠️ Skipped {total_skipped} duplicate players (already in lists or squads)"

            return (True, msg)

        except FileNotFoundError:
            return False, f"Excel file not found: {filepath}"
        except Exception as e:
            return False, f"Error loading Excel: {str(e)}"

    def _load_state_from_db(self):
        state = self.db.get_auction_state()
        self.active = bool(state.get("active", 0))
        self.paused = bool(state.get("paused", 0))
        self.current_player = state.get("current_player")
        self.current_list_index = state.get("current_list_index", 0)
        self.base_price = state.get("base_price", DEFAULT_BASE_PRICE)
        self.current_bid = state.get("current_bid", 0)
        self.highest_bidder = state.get("highest_bidder")
        self.countdown_seconds = state.get("countdown_seconds", DEFAULT_COUNTDOWN)
        self.stats_channel_id = state.get("stats_channel_id", 0)
        self.stats_message_id = state.get("stats_message_id", 0)

        db_last_bid_time = state.get("last_bid_time", 0)
        if db_last_bid_time and db_last_bid_time > 0:
            self.last_bid_time = db_last_bid_time
        else:
            self.last_bid_time = time.time()

        self.countdown_remaining = self.countdown_seconds

    def _save_state_to_db(self):
        self.db.update_auction_state(
            active=1 if self.active else 0,
            paused=1 if self.paused else 0,
            current_player=self.current_player,
            current_list_index=self.current_list_index,
            base_price=self.base_price,
            current_bid=self.current_bid,
            highest_bidder=self.highest_bidder,
            countdown_seconds=self.countdown_seconds,
            last_bid_time=self.last_bid_time,
            stats_channel_id=self.stats_channel_id,
            stats_message_id=self.stats_message_id,
        )

    @property
    def teams(self) -> Dict[str, int]:
        return self.db.get_teams()

    @property
    def team_squads(self) -> Dict[str, List[Tuple[str, int]]]:
        return self.db.get_all_squads()

    @property
    def player_lists(self) -> Dict[str, List[Tuple[str, Optional[int]]]]:
        return self.db.get_player_lists()

    @property
    def list_order(self) -> List[str]:
        return self.db.get_list_order()

    # ==================== LIST MANAGEMENT ====================

    def create_list(self, list_name: str) -> bool:
        return self.db.create_list(list_name.lower())

    def add_player_to_list(
        self, list_name: str, player: Tuple[str, Optional[int]]
    ) -> bool:
        player_name, base_price = player
        return self.db.add_player_to_list(list_name.lower(), player_name, base_price)

    def load_list_from_csv(self, list_name: str, filepath: str) -> Tuple[bool, str]:
        list_name = list_name.lower()
        try:
            players = self.file_manager.load_players_from_csv(filepath)
            if not players:
                return False, "No players found in CSV file"

            self.db.create_list(list_name)
            self.db.add_players_to_list(list_name, players)
            return True, f"Loaded {len(players)} players into {list_name}"
        except Exception as e:
            return False, str(e)

    def load_players_from_sets(
        self, num_sets: int, filepath: str = None
    ) -> Tuple[bool, str]:
        """Load the next N sets from Excel file.

        Args:
            num_sets: Number of new sets to load
            filepath: Path to Excel file (optional)
        """
        import os

        if filepath is None:
            filepath = os.path.join(os.path.dirname(__file__), DEFAULT_AUCTION_FILE)

        if not os.path.exists(filepath):
            return False, f"Excel file not found: {filepath}"

        if num_sets < 1 or num_sets > 67:
            return False, "Number of sets must be between 1 and 67"

        return self._load_auction_excel(filepath, num_sets=num_sets)

    def set_list_order(self, order: List[str]) -> Tuple[bool, str]:
        order_lower = [name.lower() for name in order]
        player_lists = self.db.get_player_lists()
        for list_name in order_lower:
            if list_name not in player_lists:
                return False, f"List '{list_name}' does not exist"
        self.db.set_list_order(order_lower)
        return True, "List order updated successfully"

    def get_list_info(self) -> str:
        return self.formatter.format_list_display(self.player_lists, self.list_order)

    def delete_set(self, set_name: str) -> Tuple[bool, str]:
        """Delete a set/list and all its players from the auction"""
        set_name_lower = set_name.lower()
        return self.db.delete_set(set_name_lower)

    def remove_players_from_list(
        self, list_name: str, player_names: List[str]
    ) -> Tuple[List[str], List[str]]:
        """Remove multiple players from a list."""
        list_name_lower = list_name.lower()
        removed = []
        not_found = []

        for player_name in player_names:
            success = self.db.remove_player_from_list(list_name_lower, player_name)
            if success:
                removed.append(player_name)
            else:
                not_found.append(player_name)

        return removed, not_found

    def skip_current_set(self) -> Tuple[int, List[str]]:
        """Skip all remaining players in current set and move to next set.
        Skipped players are moved to the 'skipped' list for potential re-auction.

        Returns:
            Tuple of (count of skipped players, list of skipped player names)
        """
        current_list = self.get_current_list_name()
        if not current_list:
            return 0, []

        # Get all unauctioned players in current set before marking
        player_lists = self.db.get_player_lists()
        players_in_set = player_lists.get(current_list, [])
        skipped_player_names = [name for name, _ in players_in_set]

        if skipped_player_names:
            # Mark all remaining unauctioned players in this set as auctioned first
            # This prevents duplicate check from failing when adding to skipped list
            skipped_count = self.db.mark_set_as_auctioned(current_list)

            # Create skipped list if doesn't exist
            skipped_list_name = "skipped"
            self.db.create_list(skipped_list_name)

            # Move players to skipped list (now they won't conflict since they're auctioned)
            for player_name, base_price in players_in_set:
                self.db.add_player_to_list(skipped_list_name, player_name, base_price)

            # Ensure skipped list is at the end of list order
            current_order = self.db.get_list_order()
            if skipped_list_name not in [o.lower() for o in current_order]:
                current_order.append(skipped_list_name)
                self.db.set_list_order(current_order)
        else:
            skipped_count = 0

        # Clear current player state
        self._reset_player_state()

        # Advance to next set
        self.current_list_index += 1
        self._save_state_to_db()

        return skipped_count, skipped_player_names

    def get_skipped_players(self) -> List[Tuple[str, Optional[int]]]:
        """Get all players in the skipped list."""
        player_lists = self.db.get_player_lists()
        return player_lists.get("skipped", [])

    # ==================== AUCTION CONTROL ====================

    def start_auction(self) -> Tuple[bool, str]:
        if self.active:
            return False, "Auction is already running"

        player_lists = self.db.get_player_lists()
        if not player_lists:
            return False, "No player lists available. Use /loadsets first."

        list_order = self.db.get_list_order()
        if not list_order:
            self.db.set_list_order(list(player_lists.keys()))

        self.active = True
        self.paused = False
        self.current_list_index = 0
        self._save_state_to_db()
        return True, "Auction started successfully"

    def stop_auction(self) -> bool:
        if not self.active:
            return False
        self._reset_state()
        return True

    def pause_auction(self) -> bool:
        if not self.active or self.paused:
            return False
        self.paused = True
        self._save_state_to_db()
        return True

    def resume_auction(self) -> bool:
        if not self.active or not self.paused:
            return False
        self.paused = False
        self.last_bid_time = time.time()  # Reset timestamp on resume
        self._save_state_to_db()
        return True

    def _reset_state(self):
        self.active = False
        self.paused = False
        self.current_player = None
        self.current_list_index = 0
        self.base_price = DEFAULT_BASE_PRICE
        self.current_bid = 0
        self.highest_bidder = None
        self.last_bid_time = time.time()
        self.stats_channel_id = 0
        self.stats_message_id = 0
        self.db.reset_auction_state()

    def _reset_player_state(self):
        self.current_player = None
        self.current_bid = 0
        self.highest_bidder = None
        self.last_bid_time = time.time()
        self.db.clear_all_auto_bids()

    # ==================== PLAYER MANAGEMENT ====================

    def reset_last_bid_time(self):
        """Resets the last bid time to now."""
        self.last_bid_time = time.time()
        self._save_state_to_db()

    def get_next_player(self) -> Tuple[bool, str, Optional[int], bool]:
        """Get the next player. Safely handles state to avoid re-auctioning active player.

        Returns:
            Tuple of (success, player_name, base_price, is_first_in_list)
        """
        # Check if current player is already active and valid
        if self.current_player:
            # Check if this player is already in a squad (sold)
            squads = self.db.get_all_squads()
            is_sold = False
            for squad in squads.values():
                for item in squad:
                    if item[0].lower() == self.current_player.lower():
                        is_sold = True
                        break
                if is_sold:
                    break

            # Also check if player was already auctioned (unsold case)
            is_auctioned = False
            player_info = self.db.find_player_by_name(self.current_player)
            if player_info:
                # player_info = (id, player_name, list_name, base_price, auctioned)
                is_auctioned = bool(player_info[4])

            if is_sold or is_auctioned:
                # Player was already processed (sold or marked unsold) - clear state and move on
                self._reset_player_state()
            else:
                # Player is active but not yet processed - continue with them
                return (True, self.current_player, self.base_price, False)

        list_order = self.db.get_list_order()
        if not list_order:
            return (False, None, None, False)

        auctioned_count = self.db.get_auctioned_count()
        is_start_of_auction = auctioned_count == 0

        old_list_index = self.current_list_index
        list_advanced = False

        # Ensure current_list_index is within bounds
        if self.current_list_index >= len(list_order):
            self.current_list_index = 0
            self._save_state_to_db()

        # First pass: try from current index to end
        start_index = self.current_list_index
        while self.current_list_index < len(list_order):
            current_list = list_order[self.current_list_index]

            if self.current_list_index > old_list_index:
                list_advanced = True

            player_data = self.db.get_random_player_from_list(current_list)
            if player_data:
                player_id, player_name, base_price = player_data
                self.db.mark_player_auctioned_by_id(player_id)

                self._reset_player_state()
                self.current_player = player_name
                self.base_price = (
                    base_price if base_price is not None else DEFAULT_BASE_PRICE
                )
                self.current_bid = self.base_price
                self._save_state_to_db()

                is_first_in_list = list_advanced or is_start_of_auction
                return (True, player_name, self.base_price, is_first_in_list)
            else:
                self.current_list_index += 1
                self._save_state_to_db()
                list_advanced = True

        # Second pass: check from beginning up to where we started (in case new sets were added)
        self.current_list_index = 0
        while self.current_list_index < start_index:
            current_list = list_order[self.current_list_index]

            player_data = self.db.get_random_player_from_list(current_list)
            if player_data:
                player_id, player_name, base_price = player_data
                self.db.mark_player_auctioned_by_id(player_id)

                self._reset_player_state()
                self.current_player = player_name
                self.base_price = (
                    base_price if base_price is not None else DEFAULT_BASE_PRICE
                )
                self.current_bid = self.base_price
                self._save_state_to_db()

                # This is first in a new list since we wrapped around
                return (True, player_name, self.base_price, True)
            else:
                self.current_list_index += 1
                self._save_state_to_db()

        return (False, None, None, False)

    def get_current_list_name(self) -> Optional[str]:
        list_order = self.db.get_list_order()
        if 0 <= self.current_list_index < len(list_order):
            return list_order[self.current_list_index]
        return None

    # ==================== ATOMIC BIDDING ====================

    async def place_bid(
        self, team: str, user_id: int, user_name: str = None, interaction_id: str = None
    ) -> BidResult:
        async with self._bid_lock:
            return await self._place_bid_internal(
                team, user_id, user_name, interaction_id, is_auto=False
            )

    async def _place_bid_internal(
        self,
        team: str,
        user_id: int,
        user_name: str = None,
        interaction_id: str = None,
        is_auto: bool = False,
    ) -> BidResult:
        from config import MAX_SQUAD_SIZE

        if not self.active:
            return BidResult(False, "Auction is not active")

        if self.paused:
            return BidResult(False, "Auction is paused - bidding not allowed")

        if not self.current_player:
            return BidResult(False, "No player is currently being auctioned")

        # CRITICAL: Check if player is already sold (prevents double-sell bug)
        squads = self.db.get_all_squads()
        for squad in squads.values():
            for item in squad:
                if item[0].lower() == self.current_player.lower():
                    return BidResult(False, "This player has already been sold")

        team_upper = team.upper()
        teams = self.db.get_teams()
        if team_upper not in teams:
            return BidResult(False, "Invalid team name")

        # CHECK SQUAD SIZE LIMIT (hard limit of 25)
        squad_count = self.db.get_squad_count(team_upper)
        if squad_count >= MAX_SQUAD_SIZE:
            return BidResult(
                False,
                f"Squad full! {team_upper} already has {MAX_SQUAD_SIZE} players.",
            )

        # CHECK OVERSEAS LIMIT (8 max) - only if current player is overseas
        is_current_player_overseas = self.db.get_player_overseas_from_list(
            self.current_player
        )
        if is_current_player_overseas:
            from config import MAX_OVERSEAS_LIMIT

            current_overseas_count = self.db.get_overseas_count(team_upper)
            if current_overseas_count >= MAX_OVERSEAS_LIMIT:
                return BidResult(
                    False,
                    f"Overseas limit reached! {team_upper} already has {MAX_OVERSEAS_LIMIT} overseas players.",
                )

        # PREVENT DOUBLE BIDDING: Check if this team is already the highest bidder
        if self.highest_bidder == team_upper:
            return BidResult(
                False,
                "You are already the highest bidder! Double bidding is not allowed.",
            )

        if self.highest_bidder is None:
            min_bid = self.base_price
        else:
            min_bid = self.current_bid + get_bid_increment(self.current_bid)

        if teams[team_upper] < min_bid:
            return BidResult(
                False,
                f"Insufficient purse. Need {format_amount(min_bid)}, have {format_amount(teams[team_upper])}",
            )

        timestamp = time.time()

        self.current_bid = min_bid
        self.highest_bidder = team_upper
        self.last_bid_time = timestamp

        self.db.record_bid(
            player_name=self.current_player,
            team_code=team_upper,
            user_id=user_id,
            user_name=user_name,
            amount=min_bid,
            timestamp=timestamp,
            is_auto_bid=is_auto,
            interaction_id=interaction_id,
        )

        self._save_state_to_db()

        auto_bids_triggered = await self._process_auto_bids(team_upper, user_id)

        return BidResult(
            success=True,
            message="Bid placed successfully",
            amount=min_bid if not auto_bids_triggered else self.current_bid,
            team=self.highest_bidder,
            is_auto_bid=is_auto,
            auto_bids_triggered=auto_bids_triggered,
            original_bid_amount=min_bid,
        )

    async def undo_last_bid(self) -> Tuple[bool, str]:
        """Undo the last bid on the current player (thread-safe)"""
        async with self._bid_lock:
            if not self.active or not self.current_player:
                return False, "No active player auction."

            # Get history
            history = self.db.get_bid_history_for_player(self.current_player)
            if not history:
                return False, "No bids to undo."

            # Remove last bid
            self.db.delete_last_bid(self.current_player)

            # Get new state
            previous = self.db.get_previous_bid(self.current_player)

            if previous:
                self.current_bid = previous["amount"]
                self.highest_bidder = previous["team_code"]
                self.last_bid_time = previous["timestamp"]
            else:
                # Back to base price
                self.current_bid = self.base_price
                self.highest_bidder = None
                self.last_bid_time = time.time()

            self._save_state_to_db()
            return (
                True,
                f"Undo successful. Current bid: {format_amount(self.current_bid)} by {self.highest_bidder or 'None'}",
            )

    # ==================== TRADING & MANUAL ====================
    def trade_player(
        self, player_name: str, from_team: str, to_team: str, price_cr: float
    ) -> Tuple[bool, str]:
        """Trade player between teams (cash trade)

        Cash Deal Logic:
        - from_team gets +price (sells player)
        - to_team pays -price (buys player)
        - Player's new salary in to_team becomes the trade price
        """
        # Convert Crores to Rupees (1 Cr = 10,000,000)
        price = int(price_cr * 10_000_000)

        # Validate target team has enough purse
        to_team_upper = to_team.upper()
        from_team_upper = from_team.upper()
        teams = self.db.get_teams()

        # Store old purses for display
        old_purse_from = teams.get(from_team_upper, 0)
        old_purse_to = teams.get(to_team_upper, 0)

        if to_team_upper in teams and teams[to_team_upper] < price:
            return (
                False,
                f"**{to_team_upper}** has insufficient purse. Need {format_amount(price)}, have {format_amount(teams[to_team_upper])}",
            )

        # Check overseas status for the message
        is_overseas = self.db.get_player_overseas_status(player_name)
        overseas_tag = " ✈️" if is_overseas else ""

        success, msg = self.db.trade_player(
            player_name, from_team_upper, to_team_upper, price
        )
        if success:
            # Update Excel after trade
            self._update_excel_after_trade()

            # Get new purses after trade
            teams = self.db.get_teams()
            new_purse_from = teams.get(from_team_upper, 0)
            new_purse_to = teams.get(to_team_upper, 0)

            # Build detailed response message
            response = f"✅ **Cash Trade Completed!**\n\n"
            response += f"**{player_name}**{overseas_tag}\n"
            response += f"**{from_team_upper}** → **{to_team_upper}** for **{format_amount(price)}**\n\n"
            response += f"**Purse Changes:**\n"
            response += f"• **{from_team_upper}**: {format_amount(old_purse_from)} → {format_amount(new_purse_from)} (+{format_amount(price)})\n"
            response += f"• **{to_team_upper}**: {format_amount(old_purse_to)} → {format_amount(new_purse_to)} (-{format_amount(price)})"

            return (True, response)
        return False, (
            msg if msg else "Trade failed. Check if player exists in source team."
        )

    def swap_players(
        self,
        player_a: str,
        team_a: str,
        player_b: str,
        team_b: str,
        compensation_cr: float = 0,
        compensation_from: str = None,
    ) -> Tuple[bool, str]:
        """
        Swap two players between teams.

        Swap Trade Logic:
        - Players exchange teams, keeping their original salaries
        - Price difference is transferred from team getting higher-valued player
        - Compensation (if any) is additional cash transfer between teams

        Example: Bumrah (18cr MI) ↔ Khaleel (4.8cr CSK)
        Difference = 18 - 4.8 = 13.2cr
        MI gets +13.2cr, CSK pays -13.2cr
        """
        team_a_upper = team_a.upper()
        team_b_upper = team_b.upper()

        # Validate teams exist
        teams = self.db.get_teams()
        if team_a_upper not in teams:
            return False, f"Invalid team: {team_a}"
        if team_b_upper not in teams:
            return False, f"Invalid team: {team_b}"

        # Store old purses for display
        old_purse_a = teams.get(team_a_upper, 0)
        old_purse_b = teams.get(team_b_upper, 0)

        # Get original player prices before swap
        original_price_a = (
            self.db.get_player_price_in_squad(team_a_upper, player_a) or 0
        )
        original_price_b = (
            self.db.get_player_price_in_squad(team_b_upper, player_b) or 0
        )

        # Check overseas status for display
        is_overseas_a = self.db.get_player_overseas_status(player_a)
        is_overseas_b = self.db.get_player_overseas_status(player_b)
        overseas_tag_a = " ✈️" if is_overseas_a else ""
        overseas_tag_b = " ✈️" if is_overseas_b else ""

        # Convert compensation to rupees
        compensation_amount = (
            int(compensation_cr * 10_000_000) if compensation_cr else 0
        )

        success, msg = self.db.swap_players(
            player_a,
            team_a_upper,
            player_b,
            team_b_upper,
            compensation_amount,
            compensation_from.upper() if compensation_from else None,
        )

        if success:
            # Update Excel after swap
            self._update_excel_after_trade()

            # Get new purses after swap
            teams = self.db.get_teams()
            new_purse_a = teams.get(team_a_upper, 0)
            new_purse_b = teams.get(team_b_upper, 0)

            # Calculate price difference for display
            price_diff = abs(original_price_a - original_price_b)

            # Build detailed response message
            response = f"✅ **Swap Trade Completed!**\n\n"
            response += f"**{player_a}**{overseas_tag_a} ({format_amount(original_price_a)}) → **{team_b_upper}**\n"
            response += f"**{player_b}**{overseas_tag_b} ({format_amount(original_price_b)}) → **{team_a_upper}**\n\n"

            # Show price difference calculation
            if price_diff > 0:
                if original_price_a > original_price_b:
                    response += f"**Price Difference:** {format_amount(original_price_a)} - {format_amount(original_price_b)} = {format_amount(price_diff)}\n"
                    response += f"({team_b_upper} pays {team_a_upper})\n\n"
                else:
                    response += f"**Price Difference:** {format_amount(original_price_b)} - {format_amount(original_price_a)} = {format_amount(price_diff)}\n"
                    response += f"({team_a_upper} pays {team_b_upper})\n\n"

            if compensation_amount > 0 and compensation_from:
                comp_from = compensation_from.upper()
                comp_to = team_b_upper if comp_from == team_a_upper else team_a_upper
                response += f"**Compensation:** {format_amount(compensation_amount)} ({comp_from} → {comp_to})\n\n"

            # Show purse changes
            change_a = new_purse_a - old_purse_a
            change_b = new_purse_b - old_purse_b
            change_str_a = (
                f"+{format_amount(change_a)}"
                if change_a >= 0
                else f"-{format_amount(abs(change_a))}"
            )
            change_str_b = (
                f"+{format_amount(change_b)}"
                if change_b >= 0
                else f"-{format_amount(abs(change_b))}"
            )

            response += f"**Purse Changes:**\n"
            response += f"• **{team_a_upper}**: {format_amount(old_purse_a)} → {format_amount(new_purse_a)} ({change_str_a})\n"
            response += f"• **{team_b_upper}**: {format_amount(old_purse_b)} → {format_amount(new_purse_b)} ({change_str_b})"

            return True, response

        return False, msg

    def _update_excel_after_trade(self):
        """Helper to update Excel file after any trade operation"""
        try:
            teams = self.db.get_teams()
            squads = self.db.get_all_squads()
            sales = self.db.get_all_sales()
            unsold = self.db.get_unsold_players_for_excel()
            released = self.db.get_released_players_for_excel()
            trades = self.db.get_all_trades()

            self.file_manager.regenerate_excel_from_db(
                self.excel_file, sales, teams, squads, unsold, released, trades
            )
        except Exception as e:
            logger.error(f"Error updating Excel after trade: {e}")

    def get_trade_log_message(self) -> str:
        """Generate a formatted trade log message for display (compact format)"""
        trades = self.db.get_all_trades()

        if not trades:
            return "📋 **Trade Log**\n_No trades yet._"

        msg = "📋 **Trade Log**\n"

        # Group swap trades to avoid duplicates (swaps create 2 records)
        seen_swaps = set()

        for trade in trades:
            trade_type = trade.get("trade_type", "cash")

            if trade_type == "swap":
                # Create a unique key for this swap pair
                swap_key = tuple(
                    sorted(
                        [
                            f"{trade['player_name']}:{trade['from_team']}",
                            f"{trade.get('swap_player', '')}:{trade['to_team']}",
                        ]
                    )
                )

                if swap_key in seen_swaps:
                    continue
                seen_swaps.add(swap_key)

                # Compact swap format
                p1 = trade["player_name"]
                p2 = trade.get("swap_player", "?")
                t1 = trade["from_team"]
                t2 = trade["to_team"]
                msg += f"🔄 {p1} ({t1})↔{p2} ({t2})"

                if (
                    trade.get("compensation_amount")
                    and trade["compensation_amount"] > 0
                ):
                    msg += f" +{format_amount(trade['compensation_amount'])}"
                msg += "\n"
            else:
                # Compact cash trade format
                p = trade["player_name"]
                t1 = trade["from_team"]
                t2 = trade["to_team"]
                price = format_amount(trade["trade_price"])
                msg += f"💵 {p}: {t1}→{t2} ({price})\n"

        return msg.strip()

    def set_trade_channel(self, channel_id: str, message_id: str = None):
        """Set the trade log channel"""
        self.db.set_trade_channel(channel_id, message_id)

    def get_trade_channel(self) -> Tuple[Optional[str], Optional[str]]:
        """Get the trade log channel and message IDs"""
        return self.db.get_trade_channel()

    def manual_add_player(
        self, team: str, player: str, price_cr: float, is_overseas: bool = False
    ) -> Tuple[bool, str]:
        """Manually add a player to a squad"""
        if price_cr <= 0:
            return False, "Price must be a positive value."

        team_upper = team.upper()
        teams = self.db.get_teams()

        if team_upper not in teams:
            return False, "Invalid team."

        # Convert Crores to Rupees (1 Cr = 10,000,000)
        price = int(price_cr * 10_000_000)

        # Check if purse would go negative
        if teams[team_upper] < price:
            return (
                False,
                f"Insufficient purse. Need {format_amount(price)}, have {format_amount(teams[team_upper])}",
            )

        if not self.db.deduct_from_purse(team_upper, price):
            return False, "Insufficient purse."

        self.db.add_to_squad(team_upper, player, price, is_overseas=is_overseas)

        # Update Excel after manual add
        try:
            teams = self.db.get_teams()
            squads = self.db.get_all_squads()
            sales = self.db.get_all_sales()
            unsold = self.db.get_unsold_players_for_excel()
            released = self.db.get_released_players_for_excel()
            trades = self.db.get_all_trades()

            self.file_manager.regenerate_excel_from_db(
                self.excel_file, sales, teams, squads, unsold, released, trades
            )
        except Exception as e:
            logger.error(f"Error updating Excel after manual add: {e}")

        return (
            True,
            f"Added **{player}** to **{team_upper}** for {format_amount(price)}",
        )

    # ==================== AUTO-BID (PROXY BIDDING) ====================

    def set_auto_bid(
        self, team: str, max_amount: int, user_id: int
    ) -> Tuple[bool, str]:
        team_upper = team.upper()
        teams = self.db.get_teams()

        if team_upper not in teams:
            return False, "Invalid team name"

        if max_amount > teams[team_upper]:
            return (
                False,
                f"Auto-bid max ({format_amount(max_amount)}) exceeds purse ({format_amount(teams[team_upper])})",
            )

        if self.current_player and max_amount < self.current_bid:
            return (
                False,
                f"Auto-bid max must be at least current bid ({format_amount(self.current_bid)})",
            )

        self.db.set_auto_bid(team_upper, max_amount, user_id)
        return True, f"Auto-bid set for {team_upper}: max {format_amount(max_amount)}"

    def clear_auto_bid(self, team: str) -> Tuple[bool, str]:
        team_upper = team.upper()
        self.db.clear_auto_bid(team_upper)
        return True, f"Auto-bid cleared for {team_upper}"

    def get_auto_bid(self, team: str) -> Optional[int]:
        return self.db.get_auto_bid(team.upper())

    async def _process_auto_bids(
        self, excluded_team: str, excluded_user: int
    ) -> List[dict]:
        auto_bids_triggered = []
        auto_bids = self.db.get_all_auto_bids()
        teams = self.db.get_teams()

        iterations = 0
        max_iterations = 100

        while iterations < max_iterations:
            iterations += 1
            next_bid = self.current_bid + get_bid_increment(self.current_bid)

            eligible = []
            for team, max_amount in auto_bids.items():
                if team == self.highest_bidder:
                    continue
                if max_amount >= next_bid and teams.get(team, 0) >= next_bid:
                    eligible.append((team, max_amount))

            if not eligible:
                break

            eligible.sort(key=lambda x: (-x[1], x[0]))
            winning_team, _ = eligible[0]

            timestamp = time.time()
            self.current_bid = next_bid
            self.highest_bidder = winning_team
            self.last_bid_time = timestamp

            self.db.record_bid(
                player_name=self.current_player,
                team_code=winning_team,
                user_id=0,
                user_name="AUTO-BID",
                amount=next_bid,
                timestamp=timestamp,
                is_auto_bid=True,
            )

            auto_bids_triggered.append({"team": winning_team, "amount": next_bid})
            teams = self.db.get_teams()

        if auto_bids_triggered:
            self._save_state_to_db()

        return auto_bids_triggered

    # ==================== RE-AUCTION UNSOLD PLAYERS ====================

    def reauction_player(self, player_name: str) -> Tuple[bool, str]:
        """Re-auction a player - moves them to 'accelerated' list at the end"""
        player_data = self.db.find_player_by_name(player_name)

        if not player_data:
            return False, f"Player **{player_name}** not found in auction database."

        player_id, actual_name, list_name, base_price, auctioned = player_data

        if auctioned == 0:
            return (
                False,
                f"**{actual_name}** is already in the auction pool (not yet auctioned).",
            )

        squads = self.db.get_all_squads()
        for team, squad in squads.items():
            for item in squad:
                if item[0].lower() == actual_name.lower():
                    return (
                        False,
                        f"**{actual_name}** was already sold to **{team}**. Use /rollback to undo the sale first.",
                    )

        # CLEANUP: Safely remove previous sales record (including UNSOLD status)
        try:
            self.db.delete_sale(actual_name)
        except Exception as e:
            logger.error(f"Error deleting previous sale record for reauction: {e}")

        if base_price is None:
            base_price = DEFAULT_BASE_PRICE

        # Create "accelerated" list if it doesn't exist
        accelerated_list_name = "accelerated"
        self.db.create_list(accelerated_list_name)

        # Move player to accelerated list (updates list_name and resets auctioned status)
        self.db.move_player_to_list_by_id(player_id, accelerated_list_name)

        # Ensure "accelerated" list is at the end of list order
        current_order = self.db.get_list_order()
        if accelerated_list_name not in [o.lower() for o in current_order]:
            current_order.append(accelerated_list_name)
            self.db.set_list_order(current_order)

        return (
            True,
            f"**{actual_name}** has been added to **Accelerated** list with base price {format_amount(base_price)}",
        )

    # ==================== RELEASE PLAYERS ====================

    def release_retained_player(
        self, team: str, player_name: str, before_auction: bool = False
    ) -> Tuple[bool, str]:
        """Release any player from a squad back into auction (Retained or Sold)"""
        team_upper = team.upper()
        teams = self.db.get_teams()

        if team_upper not in teams:
            return False, "Invalid team name"

        # Check squad first (covers both retained and bought players)
        squads = self.db.get_all_squads()
        if team_upper not in squads:
            return False, f"{team_upper} has no players to release."

        player_found = None
        for item in squads[team_upper]:
            p_name = item[0]
            price = item[1]
            if p_name.lower() == player_name.lower():
                player_found = (p_name, price)
                break

        if not player_found:
            return False, f'"{player_name}" is not in {team_upper}\'s squad.'

        p_name, salary = player_found

        # Check if player is overseas before removing from squad
        is_overseas = self.db.get_player_overseas_status(p_name)

        # Refund logic
        current_purse = teams[team_upper]
        new_purse = current_purse + salary

        # 1. Update Purse
        self.db.update_team_purse(team_upper, new_purse)

        # 2. Remove from Squad and Sales using DB methods
        try:
            self.db.remove_from_squad(team_upper, player_name)
            self.db.delete_sale(player_name)
        except Exception as e:
            logger.error(f"Error releasing player from DB: {e}")

        # 3. Add released player to sales table as RELEASED (not UNSOLD) with overseas emoji
        try:
            # Add ✈️ emoji to released player name if overseas
            display_name = (
                f"{p_name} ✈️" if is_overseas and "✈️" not in p_name else p_name
            )
            self.db.record_sale(
                f"{display_name} (RELEASED from {team_upper})",
                "RELEASED",
                salary,
                total_bids=0,
            )
        except Exception as e:
            logger.error(f"Error adding released player to sales: {e}")

        # 4. Add to "released" list for re-auction (at the START of list order)
        released_list_name = "released"
        self.db.create_list(released_list_name)

        # IMPORTANT: per request, set base price of released players to 2 Cr
        two_cr = 2 * 10_000_000  # 2 Crore in rupees
        # Add player with overseas status preserved
        self.db.add_player_to_list_with_overseas_flag(
            released_list_name, p_name, two_cr, is_overseas
        )

        # Ensure "released" list is at the START of list order
        current_order = self.db.get_list_order()
        if released_list_name not in [o.lower() for o in current_order]:
            # Insert at beginning
            current_order.insert(0, released_list_name)
            self.db.set_list_order(current_order)
        elif current_order[0].lower() != released_list_name:
            # Move to beginning if not already there
            current_order = [
                o for o in current_order if o.lower() != released_list_name
            ]
            current_order.insert(0, released_list_name)
            self.db.set_list_order(current_order)

        # 5. Update Excel
        try:
            teams = self.db.get_teams()
            squads = self.db.get_all_squads()
            sales = self.db.get_all_sales()
            unsold = self.db.get_unsold_players_for_excel()
            released = self.db.get_released_players_for_excel()
            trades = self.db.get_all_trades()

            self.file_manager.regenerate_excel_from_db(
                self.excel_file, sales, teams, squads, unsold, released, trades
            )
        except Exception as e:
            logger.error(f"Error updating Excel after release: {e}")

        return (
            True,
            f'Released {p_name} from {team_upper}. Refunded {format_amount(salary)}. Player added to "Released" set (will appear first). Base price set to 2Cr.',
        )

    # ==================== SALE FINALIZATION ====================

    async def finalize_sale(self) -> Tuple[bool, Optional[str], int]:
        """Finalize the current player sale using atomic database operations."""
        async with self._bid_lock:
            self._load_state_from_db()

            if not self.current_player:
                return False, None, 0

            player = self.current_player

            # CRITICAL FIX: Get the winning bid from bid_history table - authoritative source
            highest_bid = self.db.get_highest_bid_for_player(player)

            if highest_bid:
                team = highest_bid["team_code"]
                amount = highest_bid["amount"]
                bid_count = self.db.count_bids_for_player(player)

                # Try to get overseas status from player_lists (if loaded)
                # Note: Auction Excel doesn't have overseas data - admin must track manually
                is_overseas = self.db.get_player_overseas_from_list(player)

                # Use atomic sale operation - all or nothing
                success, error_msg = self.db.finalize_sale_atomic(
                    player, team, amount, bid_count, is_overseas
                )

                if not success:
                    logger.warning(f"Atomic sale failed for {player}: {error_msg}")
                    # Player might already be sold - clear state and move on
                    self._reset_player_state()
                    self._save_state_to_db()
                    return False, None, 0

                # Update Excel (non-critical, outside transaction)
                try:
                    teams = self.db.get_teams()
                    squads = self.db.get_all_squads()
                    sales = self.db.get_all_sales()
                    unsold = self.db.get_unsold_players_for_excel()
                    released = self.db.get_released_players_for_excel()
                    trades = self.db.get_all_trades()

                    await asyncio.to_thread(
                        self.file_manager.regenerate_excel_from_db,
                        self.excel_file,
                        sales,
                        teams,
                        squads,
                        unsold,
                        released,
                        trades,
                    )
                except Exception as e:
                    logger.error(f"Error saving to Excel: {e}")

                # Clear player state AFTER successful sale
                self._reset_player_state()
                self._save_state_to_db()

                return True, team, amount

            else:
                # No bids - player goes UNSOLD
                team = "UNSOLD"
                amount = self.base_price or 0

                # Use atomic unsold recording
                try:
                    self.db.record_unsold_atomic(player, amount)
                except Exception as e:
                    logger.error(f"Failed to record UNSOLD in DB: {e}")

                # Update Excel (non-critical)
                try:
                    teams = self.db.get_teams()
                    squads = self.db.get_all_squads()
                    sales = self.db.get_all_sales()
                    unsold = self.db.get_unsold_players_for_excel()
                    released = self.db.get_released_players_for_excel()
                    trades = self.db.get_all_trades()

                    await asyncio.to_thread(
                        self.file_manager.regenerate_excel_from_db,
                        self.excel_file,
                        sales,
                        teams,
                        squads,
                        unsold,
                        released,
                        trades,
                    )
                except Exception as e:
                    logger.error(f"Error saving to Excel (Unsold): {e}")

                # Clear state
                self._reset_player_state()
                self._save_state_to_db()

                return True, "UNSOLD", amount

    # ==================== ADMIN OPERATIONS ====================

    def set_countdown(self, seconds: int) -> bool:
        if seconds < 5 or seconds > 300:
            return False
        self.countdown_seconds = seconds
        self._save_state_to_db()
        return True

    def set_team_purse(self, team: str, amount: int) -> bool:
        team_upper = team.upper()
        if amount < 0:
            return False
        return self.db.update_team_purse(team_upper, amount)

    def set_stats_channel(self, channel_id: int):
        self.stats_channel_id = channel_id
        self._save_state_to_db()

    def set_countdown_gap(self, seconds: int):
        """Set the time gap between last bid and start of countdown."""
        self.countdown_gap = seconds

    def set_player_gap(self, seconds: int):
        """Set the gap between players in auction."""
        self.player_gap = seconds

    def _create_backup(self) -> str:
        """Create a JSON backup of all auction data."""
        import json
        from datetime import datetime

        backup_data = {
            "timestamp": datetime.now().isoformat(),
            "teams": self.db.get_teams(),
            "squads": {
                team: [
                    (item[0], item[1], item[2] if len(item) > 2 else False)
                    for item in players
                ]
                for team, players in self.db.get_all_squads().items()
            },
            "sales": self.db.get_all_sales(),
            "player_lists": self.db.get_all_lists(),
            "list_order": self.db.get_list_order(),
            "state": {
                "active": self.active,
                "paused": self.paused,
                "current_player": self.current_player,
                "current_bid": self.current_bid,
                "highest_bidder": self.highest_bidder,
                "base_price": self.base_price,
            },
        }

        backup_filename = (
            f"auction_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        with open(backup_filename, "w", encoding="utf-8") as f:
            json.dump(backup_data, f, indent=2, default=str)

        return backup_filename

    def rollback_last_sale(self) -> Optional[dict]:
        sale = self.db.rollback_last_sale()

        if sale:
            try:
                teams = self.db.get_teams()
                squads = self.db.get_all_squads()
                sales = self.db.get_all_sales()
                unsold = self.db.get_unsold_players_for_excel()
                released = self.db.get_released_players_for_excel()
                trades = self.db.get_all_trades()

                self.file_manager.regenerate_excel_from_db(
                    self.excel_file, sales, teams, squads, unsold, released, trades
                )
            except Exception as e:
                logger.error(f"Error updating Excel on rollback: {e}")

        return sale

    def clear_all_data(self, create_backup: bool = True) -> Optional[str]:
        """
        Clear auction progress (buys, bids, releases) but KEEP retained players.
        Also clears player_lists so sets can be reloaded.
        Resets teams to config purse values.

        Args:
            create_backup: Whether to create a backup before clearing (default: True)

        Returns:
            Path to backup file if created, None otherwise
        """
        # Create backup before destructive operation (if requested)
        backup_path = None
        if create_backup:
            try:
                backup_path = self._create_backup()
                logger.info(f"Created backup before clear: {backup_path}")
            except Exception as e:
                logger.warning(f"Failed to create backup before clear: {e}")

        # 1. Clear State
        self._reset_state()

        # 2. Clear Database Tables (granularly)
        self.db.clear_auction_buys()  # Deletes bought/traded players, keeps retained
        self.db.clear_released_players()  # Deletes released players list
        self.db.clear_unsold_players()  # Deletes accelerated/unsold players list
        self.db.clear_sales()  # Deletes sales history
        self.db.clear_bid_history()  # Deletes bid history
        self.db.clear_trade_history()  # Deletes trade history
        self.db.clear_all_auto_bids()  # Deletes auto bids

        # Clear player_lists so sets can be reloaded from scratch
        self.db.clear_player_lists()

        # Also clear skipped players list (already included in clear_player_lists)
        # try:
        #     self.db.delete_set("skipped")
        # except Exception:
        #     pass

        # Reset max loaded set counter so /loadsets starts fresh
        self.db.set_max_loaded_set(0)

        # 3. Reset Team Purses to config values
        # The TEAMS config already has the remaining purse after retained players.
        # We should NOT deduct retained cost again - it's already accounted for.
        from config import TEAMS

        for team_code, purse in TEAMS.items():
            self.db.update_team_purse(team_code, purse)
            # Also update original_purse to match config
            self.db.set_original_purse(team_code, purse)

        # 4. Regenerate Excel with clean state (retained players only)
        try:
            teams = self.db.get_teams()
            squads = self.db.get_all_squads()  # Will contain only retained players now
            self.file_manager.initialize_excel_with_retained_players(
                self.excel_file, teams, squads
            )
        except Exception as e:
            logger.error(f"Error reinitializing Excel after clear: {e}")

        return backup_path

    # ==================== DISPLAY HELPERS ====================

    def get_purse_display(self) -> str:
        return self.formatter.format_purse_display(self.db.get_teams(), self.db)

    def get_bid_history_display(self, player: str = None, limit: int = 5) -> str:
        if player:
            bids = self.db.get_bid_history_for_player(player)
        else:
            bids = self.db.get_recent_bids(limit)

        if not bids:
            return "No bids recorded."

        msg = "**Recent Bids:**\n```\n"
        for bid in bids[-limit:]:
            auto = "[AUTO]" if bid.get("is_auto_bid") else ""
            player_name = bid.get("player_name", "")
            player_suffix = (
                f" ({player_name[:15]})" if player_name and not player else ""
            )
            msg += f"{bid['team_code']:6} : {format_amount(bid['amount'])} {auto}{player_suffix}\n"
        msg += "```"
        return msg

    def get_team_bid_history_display(self, team: str, limit: int = 20) -> str:
        """Get bid history summary for a specific team - shows entry/exit/win per player"""
        team_upper = team.upper()
        summary = self.db.get_team_bid_summary(team_upper, limit)

        if not summary:
            return f"No bid history found for **{team_upper}**."

        msg = f"**{team_upper} Bidding Summary (Last {min(limit, len(summary))} players):**\n```\n"
        msg += f"{'Player':<22} {'Entry':>8} {'Exit/Won':>10} {'Result':<8}\n"
        msg += "=" * 52 + "\n"

        for item in summary:
            player = item.get("player_name", "Unknown")[:21]
            entry = format_amount(item.get("first_bid", 0))
            exit_bid = format_amount(item.get("last_bid", 0))
            won = item.get("won", False)
            result = "✅ WON" if won else "❌ Exit"
            msg += f"{player:<22} {entry:>8} {exit_bid:>10} {result:<8}\n"

        msg += "```"
        return msg

    def get_stats_message(self) -> str:
        data = self.db.get_stats_data()

        msg = "**📊 LIVE AUCTION STATS**\n\n"

        if data["most_expensive"]:
            # Format: Player Name - Price (Team Name)
            p_name = data["most_expensive"]["player_name"]
            price = format_amount(data["most_expensive"]["final_price"])
            t_code = data["most_expensive"]["team_code"]
            msg += f"💰 **Most Expensive:** {p_name} - {price} ({t_code})\n"
        else:
            msg += "💰 **Most Expensive:** None\n"

        if data["most_players"]:
            msg += f"📈 **Most Players:** {data['most_players']['team_code']} ({data['most_players']['cnt']})\n"
        else:
            msg += "📉 **Most Players:** None\n"  # Fixed label

        if data["least_players"]:
            msg += f"📉 **Least Players:** {data['least_players']['team_code']} ({data['least_players']['cnt']})\n"
        else:
            msg += "📉 **Least Players:** None\n"

        msg += "\n**💳 Team Purses:**\n```\n"
        for team, purse in data["purses"].items():
            msg += f"{team:6}: {format_amount(purse)}\n"
        msg += "```"

        return msg

    def get_status_display(self) -> str:
        status = "**Auction Status:**\n"
        status += f"Active: {'Yes' if self.active else 'No'}\n"
        status += f"Paused: {'Yes' if self.paused else 'No'}\n"

        if self.current_player:
            status += f"Current Player: **{self.current_player}**\n"
            status += f"Base Price: {format_amount(self.base_price)}\n"
            status += f"Current Bid: {format_amount(self.current_bid)}\n"
            status += f"Highest Bidder: **{self.highest_bidder or 'None'}**\n"

            import time as time_module

            elapsed = int(time_module.time() - self.last_bid_time)
            status += f"Time since last bid: {elapsed}s\n"

        current_list = self.get_current_list_name()
        if current_list:
            status += f"Current List: **{current_list}**\n"

        auto_bids = self.db.get_all_auto_bids()
        if auto_bids:
            status += f"Auto-bids Active: {', '.join(auto_bids.keys())}\n"

        return status

    # ==================== NEW: BASE PRICE CHANGE ====================
    def change_base_price(self, players_arg: str, price_cr: float) -> Tuple[bool, str]:
        """
        Change base price for players.
        players_arg: comma-separated player names OR the keyword 'released' to change all released players.
        price_cr: price in Crores (float)
        """
        new_price = int(price_cr * 10_000_000)
        players_arg = players_arg.strip()
        if not players_arg:
            return False, "Please provide player names or 'released'"

        if players_arg.lower() == "released":
            updated = self.db.change_base_price_for_list("released", new_price)
            # Regenerate excel
            try:
                teams = self.db.get_teams()
                squads = self.db.get_all_squads()
                sales = self.db.get_all_sales()
                unsold = self.db.get_unsold_players_for_excel()
                released = self.db.get_released_players_for_excel()
                trades = self.db.get_all_trades()

                self.file_manager.regenerate_excel_from_db(
                    self.excel_file, sales, teams, squads, unsold, released, trades
                )
            except Exception as e:
                logger.error(f"Error updating Excel after change base price: {e}")

            return (
                True,
                f"Updated base price for {updated} players in RELEASED set to {format_amount(new_price)}",
            )
        else:
            names = [n.strip() for n in players_arg.split(",") if n.strip()]
            if not names:
                return False, "No valid player names provided."

            updated, not_found = self.db.change_base_price_for_players(names, new_price)

            try:
                teams = self.db.get_teams()
                squads = self.db.get_all_squads()
                sales = self.db.get_all_sales()
                unsold = self.db.get_unsold_players_for_excel()
                released = self.db.get_released_players_for_excel()
                trades = self.db.get_all_trades()

                self.file_manager.regenerate_excel_from_db(
                    self.excel_file, sales, teams, squads, unsold, released, trades
                )
            except Exception as e:
                logger.error(f"Error updating Excel after change base price: {e}")

            msg = f"Updated base price for {updated} matching entries to {format_amount(new_price)}."
            if not_found:
                msg += f" Not found in player_lists: {', '.join(not_found)}"
            return True, msg

    # ==================== NEW: FIND PLAYER (SEARCH) ====================
    def find_player(self, query: str) -> str:
        """
        Search across player_lists (unauctioned), team_squads, and sales for query substring (case-insensitive).
        Returns a formatted string summarizing matches.
        """
        q = query.strip().lower()
        if not q:
            return "Please provide a player name to search."

        results = []

        # 1) Search unauctioned player_lists
        all_lists = self.db.get_all_lists()
        list_matches = []
        for list_name, players in all_lists.items():
            for p in players:
                if q in p["player_name"].lower():
                    price = p["base_price"]
                    list_matches.append(
                        (p["player_name"], list_name, price if price else None)
                    )

        # 2) Search squads
        squads = self.db.get_all_squads()
        squad_matches = []
        for team, players in squads.items():
            for item in players:
                pname = item[0]
                price = item[1]
                if q in pname.lower():
                    squad_matches.append((pname, team, price))

        # 3) Search sales
        sales = self.db.get_all_sales()
        sales_matches = []
        for s in sales:
            pname = s.get("player_name", "")
            if q in pname.lower():
                sales_matches.append(
                    (pname, s.get("team_code", ""), s.get("final_price", 0))
                )

        # Build message
        msg = f"**Search results for '{query}':**\n\n"

        if list_matches:
            msg += "📋 Unauctioned Lists:\n```\n"
            for pname, lname, price in list_matches:
                price_str = format_amount(price) if price else "N/A"
                # Check if overseas
                is_overseas = self.db.get_player_overseas_from_list(pname)
                overseas_marker = " ✈️" if is_overseas else ""
                msg += f"{pname:30}{overseas_marker} | Set: {lname.upper():8} | Base: {price_str}\n"
            msg += "```\n"
        else:
            msg += "📋 Unauctioned Lists: None\n"

        if squad_matches:
            msg += "🟦 Squad (Owned Players):\n```\n"
            for pname, team, price in squad_matches:
                # Check if overseas from squad
                is_overseas = self.db.get_player_overseas_status(pname)
                overseas_marker = " ✈️" if is_overseas else ""
                msg += f"{pname:30}{overseas_marker} | Team: {team:6} | Salary: {format_amount(price)}\n"
            msg += "```\n"
        else:
            msg += "🟦 Squad (Owned Players): None\n"

        if sales_matches:
            msg += "🏷️ Sales / Released / Unsold:\n```\n"
            for pname, tcode, price in sales_matches:
                price_str = format_amount(price) if price else "N/A"
                msg += f"{pname:30} | Team/Sale: {tcode:8} | Price: {price_str}\n"
            msg += "```\n"
        else:
            msg += "🏷️ Sales / Released / Unsold: None\n"

        return msg.strip()


class AuctionState:
    pass
