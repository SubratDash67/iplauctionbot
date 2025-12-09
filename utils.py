"""
Utility functions for Discord Auction Bot
Contains helper functions for formatting, file operations, etc.
"""

import csv
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Set
import os

# Set up module-level logger
logger = logging.getLogger(__name__)


def sanitize_csv_value(value: str) -> str:
    """Sanitize CSV values to prevent formula injection attacks.

    Excel/Sheets can execute formulas starting with =, +, -, @, tab, or carriage return.
    This function prefixes such values with a single quote to prevent execution.
    """
    if not value:
        return value

    dangerous_chars = ("=", "+", "-", "@", "\t", "\r")
    if value.startswith(dangerous_chars):
        return f"'{value}"
    return value


# -----------------------------------------------------------
#  AMOUNT FORMATTER  → Converts Rs amounts to cr / lakh (readable)
# -----------------------------------------------------------
def format_amount(num: Optional[int]) -> str:
    """Convert numeric rupee amount to readable cr/lakh format.

    Rules:
      - >= 1,00,00,000 -> show in crores with up to 2 decimals, suffixed "cr"
      - >= 1,00,000 -> show in lakhs with up to 2 decimals, suffixed "L"
      - None or small -> formatted number
    """
    if num is None:
        return "0"

    try:
        n = int(num)
    except Exception:
        return str(num)

    if n >= 10_000_000:
        val = n / 10_000_000.0
        s = f"{val:.2f}".rstrip("0").rstrip(".")
        return f"{s}cr"
    if n >= 100_000:
        val = n / 100_000.0
        s = f"{val:.2f}".rstrip("0").rstrip(".")
        return f"{s}L"
    return f"{n:,}"


def _save_workbook_with_retry(
    wb, filepath: str, max_retries: int = 3, delay: float = 0.5
) -> None:
    """Save workbook with retry logic for file lock issues.

    Args:
        wb: openpyxl Workbook object
        filepath: Path to save to
        max_retries: Maximum number of retry attempts
        delay: Delay between retries in seconds

    Raises:
        PermissionError: If file is locked after all retries
        Exception: For other save errors
    """
    import time as time_module

    last_error = None
    for attempt in range(max_retries):
        try:
            wb.save(filepath)
            return
        except PermissionError as e:
            last_error = e
            if attempt < max_retries - 1:
                logger.warning(
                    f"Excel file locked, retrying in {delay}s... (attempt {attempt + 1}/{max_retries})"
                )
                time_module.sleep(delay)
            continue
        except Exception as e:
            raise e

    logger.error(
        f"Failed to save Excel file after {max_retries} attempts: {last_error}"
    )
    raise PermissionError(
        f"Excel file '{filepath}' is locked by another process. Please close it and try again."
    )


class FileManager:
    """Handles file operations for the auction bot"""

    @staticmethod
    def load_players_from_csv(filepath: str) -> List[Tuple[str, Optional[int]]]:
        """
        Load players from a CSV file.
        Returns list of (player_name, base_price_or_None).
        Supports:
          - simple single-column list (Name)
          - two-column Name,BasePrice
          - IPL-like CSV where base price is present in a numeric column
        """
        players: List[Tuple[str, Optional[int]]] = []
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = [r for r in reader if any(cell.strip() for cell in r)]
                if not rows:
                    return players

                # Try detect header with BasePrice column
                first_row = rows[0]
                header = [c.strip().lower() for c in first_row]

                # If header contains 'player' or 'name', and 'price' or 'base' try parse with header
                has_name = any("name" in c for c in header)
                has_price = any("price" in c or "base" in c for c in header)

                start_index = 1 if has_name else 0

                if has_name and has_price:
                    # Use headers to find columns
                    name_idx = next((i for i, c in enumerate(header) if "name" in c), 0)
                    price_idx = next(
                        (
                            i
                            for i, c in enumerate(header)
                            if ("price" in c or "base" in c)
                        ),
                        None,
                    )
                    for row in rows[start_index:]:
                        try:
                            name = row[name_idx].strip()
                        except IndexError:
                            continue
                        price = None
                        if price_idx is not None and price_idx < len(row):
                            raw = (
                                row[price_idx]
                                .replace(",", "")
                                .replace("Rs", "")
                                .replace("rs", "")
                                .strip()
                            )
                            if raw and (raw.replace(".", "", 1).isdigit()):
                                try:
                                    price = int(float(raw))
                                except (ValueError, TypeError):
                                    price = None
                        if name and not name.isdigit():
                            players.append((name, price))
                else:
                    # No clear header: guess formats per row
                    for row in rows:
                        # prefer rows with at least one non-empty cell
                        if not row:
                            continue
                        # name is first non-empty
                        name = None
                        price = None
                        for i, cell in enumerate(row):
                            if cell and cell.strip():
                                name = cell.strip()
                                # look ahead for price in next few cols
                                for next_cell in row[i + 1 : i + 4]:
                                    if next_cell and next_cell.strip():
                                        raw = (
                                            next_cell.replace(",", "")
                                            .replace("Rs", "")
                                            .replace("rs", "")
                                            .strip()
                                        )
                                        if raw and (raw.replace(".", "", 1).isdigit()):
                                            try:
                                                price = int(float(raw))
                                            except (ValueError, TypeError):
                                                price = None
                                            break
                                break
                        if name and not name.isdigit():
                            players.append((name, price))
            return players
        except FileNotFoundError:
            logger.error(f"CSV file not found: {filepath}")
            raise FileNotFoundError(f"CSV file not found: {filepath}")
        except csv.Error as e:
            logger.error(f"CSV parsing error in {filepath}: {e}")
            raise ValueError(f"Error parsing CSV file: {str(e)}")
        except UnicodeDecodeError as e:
            logger.error(f"Encoding error in {filepath}: {e}")
            raise ValueError(f"File encoding error. Please use UTF-8: {str(e)}")
        except Exception as e:
            logger.error(f"Unexpected error reading CSV {filepath}: {e}")
            raise Exception(f"Error reading CSV file: {str(e)}")

    @staticmethod
    def initialize_excel(filepath: str) -> None:
        """Initialize Excel file with headers - creates all required sheets"""
        wb = openpyxl.Workbook()

        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        trade_header_fill = PatternFill(
            start_color="7030A0", end_color="7030A0", fill_type="solid"
        )

        # Sheet 1: Auction Results
        sheet = wb.active
        sheet.title = "Auction Results"
        headers = ["Player Name", "Team", "Final Price", "Timestamp"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Sheet 2: Team Summary
        team_sheet = wb.create_sheet("Team Summary")
        team_sheet.append(
            ["Team", "Players", "Overseas", "Total Spent", "Remaining Purse"]
        )
        for cell in team_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Sheet 3: Unsold Players
        unsold_sheet = wb.create_sheet("Unsold Players")
        unsold_sheet.append(["Player", "Set Name", "Base Price"])
        for cell in unsold_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Sheet 4: Trade History
        trade_sheet = wb.create_sheet("Trade History")
        trade_sheet.append(
            [
                "Player Name",
                "From Team",
                "To Team",
                "Trade Price",
                "Original Price",
                "Timestamp",
            ]
        )
        for cell in trade_sheet[1]:
            cell.fill = trade_header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Sheet 5+: Individual Team Sheets
        from config import TEAMS

        for team_code in sorted(TEAMS.keys()):
            team_individual_sheet = wb.create_sheet(team_code)
            team_individual_sheet.append(["Player Name", "Price", "Type"])
            for cell in team_individual_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        wb.save(filepath)

    @staticmethod
    def initialize_excel_with_retained_players(
        filepath: str,
        teams: Dict[str, int],
        squad_data: Dict[str, List[Tuple[str, int, bool]]],
    ) -> None:
        """Initialize Excel file with squad data (usually retained players) in team sheets.

        Args:
            filepath: Path to Excel file
            teams: Dict of team purses
            squad_data: Dict mapping team_code to list of (player_name, price, is_overseas) tuples
        """
        try:
            from config import TEAMS as TEAM_CONFIG

            # Initialize fresh Excel
            FileManager.initialize_excel(filepath)
            wb = openpyxl.load_workbook(filepath)

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            summary_fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid"
            )
            summary_font = Font(bold=True)

            # Update each team sheet with their squad data
            for team_code in sorted(TEAM_CONFIG.keys()):
                if team_code in wb.sheetnames:
                    del wb[team_code]

                ts = wb.create_sheet(team_code)
                ts.append(["Player Name", "Price", "Type"])
                for cell in ts[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                squad = squad_data.get(team_code, [])
                total_spent = 0
                overseas_count = 0
                for item in squad:
                    # squad_data format: (player_name, price, is_overseas)
                    player_name = item[0]
                    price = item[1]
                    is_overseas = item[2] if len(item) > 2 else False

                    display_name = f"✈️ {player_name}" if is_overseas else player_name
                    safe_player = sanitize_csv_value(display_name)
                    ts.append([safe_player, format_amount(price), "Retained"])
                    total_spent += price
                    if is_overseas:
                        overseas_count += 1

                # Summary rows
                ts.append([])
                purse_left = teams.get(team_code, 0)

                summary_row = ts.max_row + 1
                ts.append(["Total Players", len(squad), ""])
                ts.append(["Overseas Players", overseas_count, ""])
                ts.append(["Total Spent", format_amount(total_spent), ""])
                ts.append(["Purse Remaining", format_amount(purse_left), ""])

                for row_num in range(summary_row, ts.max_row + 1):
                    for cell in ts[row_num]:
                        cell.fill = summary_fill
                        cell.font = summary_font

                ts.column_dimensions["A"].width = 30
                ts.column_dimensions["B"].width = 15
                ts.column_dimensions["C"].width = 12

            # Also update Team Summary
            if "Team Summary" in wb.sheetnames:
                del wb["Team Summary"]
            summary_sheet = wb.create_sheet("Team Summary", 1)
            summary_sheet.append(
                ["Team", "Players", "Overseas", "Total Spent", "Remaining Purse"]
            )
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for team_code in sorted(TEAM_CONFIG.keys()):
                squad = squad_data.get(team_code, [])
                total_spent = 0
                overseas_count = 0
                for item in squad:
                    total_spent += item[1]
                    if len(item) > 2 and item[2]:
                        overseas_count += 1
                remaining = teams.get(team_code, 0)
                summary_sheet.append(
                    [
                        team_code,
                        len(squad),
                        overseas_count,
                        format_amount(total_spent),
                        format_amount(remaining),
                    ]
                )

            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error initializing Excel with retained players: {str(e)}")

    @staticmethod
    def save_player_to_excel(
        filepath: str,
        player: str,
        team: str,
        price: int,
        remaining_purse: int,
        is_overseas: bool = False,
    ) -> None:
        """Save a sold player to the Excel file"""
        try:
            if not os.path.exists(filepath):
                FileManager.initialize_excel(filepath)
            wb = openpyxl.load_workbook(filepath)
            sheet = wb["Auction Results"]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # Format price as cr/L for readability
            formatted_price = format_amount(price)
            # Add ✈️ emoji for overseas players
            display_name = (
                f"{player} ✈️" if is_overseas and "✈️" not in player else player
            )
            # Sanitize player name to prevent formula injection
            safe_player = sanitize_csv_value(display_name)
            sheet.append([safe_player, team, formatted_price, timestamp])
            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error saving to Excel: {str(e)}")

    @staticmethod
    def update_team_summary(
        filepath: str,
        teams: Dict[str, int],
        team_squads: Dict[str, List[Tuple[str, int, bool]]],
    ) -> None:
        """Update the team summary sheet with current data"""
        try:
            if not os.path.exists(filepath):
                FileManager.initialize_excel(filepath)
            wb = openpyxl.load_workbook(filepath)
            if "Team Summary" in wb.sheetnames:
                del wb["Team Summary"]
            ts = wb.create_sheet("Team Summary", 1)  # Position after Auction Results
            ts.append(["Team", "Players", "Overseas", "Total Spent", "Remaining Purse"])
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ts[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            from config import TEAMS

            for team in sorted(teams.keys()):
                purse_left = teams.get(team, 0)
                squad = team_squads.get(team, [])
                # Handle both old and new formats
                spent = 0
                overseas_count = 0
                for item in squad:
                    if len(item) >= 3:
                        spent += item[1]
                        if item[2]:
                            overseas_count += 1
                    else:
                        spent += item[1]
                # Format amounts as cr/L for readability
                ts.append(
                    [
                        team,
                        len(squad),
                        overseas_count,
                        format_amount(spent),
                        format_amount(purse_left),
                    ]
                )
            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error updating team summary: {str(e)}")

    @staticmethod
    def update_unsold_players_sheet(
        filepath: str,
        unsold_players: List[Tuple[str, str, Optional[int]]],
    ) -> None:
        """Update the unsold players sheet

        Args:
            filepath: Path to Excel file
            unsold_players: List of (player_name, set_name, base_price)
        """
        try:
            if not os.path.exists(filepath):
                FileManager.initialize_excel(filepath)
            wb = openpyxl.load_workbook(filepath)

            # Remove existing sheet if present
            if "Unsold Players" in wb.sheetnames:
                del wb["Unsold Players"]

            # Create at position 2 (after Team Summary)
            us = wb.create_sheet("Unsold Players", 2)
            us.append(["Player", "Set Name", "Base Price"])

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            for cell in us[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for player_name, set_name, base_price in unsold_players:
                # Add ✈️ emoji if not already present and player is overseas
                display_name = player_name
                if "✈️" not in player_name:
                    # Check if player name originally had emoji (it might have been removed)
                    # For now, preserve as-is since unsold players should retain their original names
                    pass
                safe_player = sanitize_csv_value(display_name)
                us.append(
                    [
                        safe_player,
                        set_name.upper() if set_name else "N/A",
                        format_amount(base_price) if base_price else "N/A",
                    ]
                )

            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error updating unsold players sheet: {str(e)}")

    @staticmethod
    def update_released_players_sheet(
        filepath: str,
        released_players: List[Tuple[str, str, Optional[int]]],
    ) -> None:
        """Update the released players sheet

        Args:
            filepath: Path to Excel file
            released_players: List of (player_name, released_from, price)
        """
        try:
            if not os.path.exists(filepath):
                FileManager.initialize_excel(filepath)
            wb = openpyxl.load_workbook(filepath)

            # Remove existing sheet if present
            if "Released Players" in wb.sheetnames:
                del wb["Released Players"]

            # Create at position 3 (after Team Summary and Unsold Players)
            rp = wb.create_sheet("Released Players", 3)
            rp.append(["Player", "Released Info", "Price"])

            header_fill = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            for cell in rp[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for player_name, info, price in released_players:
                # Player names from sales table should already have ✈️ if they were overseas
                # Just preserve the display name as-is
                display_name = player_name
                safe_player = sanitize_csv_value(display_name)
                rp.append(
                    [
                        safe_player,
                        info if info else "N/A",
                        format_amount(price) if price else "N/A",
                    ]
                )

            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error updating released players sheet: {str(e)}")

    @staticmethod
    def update_trade_history_sheet(
        filepath: str,
        trades: List[dict],
    ) -> None:
        """Update the trade history sheet

        Args:
            filepath: Path to Excel file
            trades: List of trade records with keys: player_name, from_team, to_team, trade_price,
                    original_price, traded_at, trade_type, swap_player, swap_player_price,
                    compensation_amount, compensation_direction
        """
        try:
            if not os.path.exists(filepath):
                FileManager.initialize_excel(filepath)
            wb = openpyxl.load_workbook(filepath)

            # Remove existing sheet if present
            if "Trade History" in wb.sheetnames:
                del wb["Trade History"]

            # Create at position 3 (after Unsold Players)
            ts = wb.create_sheet("Trade History", 3)
            ts.append(
                [
                    "Player Name",
                    "From Team",
                    "To Team",
                    "Trade Price/Salary",
                    "Original Price",
                    "Trade Type",
                    "Swap Player",
                    "Swap Player Salary",
                    "Compensation",
                    "Timestamp",
                ]
            )

            header_fill = PatternFill(
                start_color="7030A0", end_color="7030A0", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ts[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            if trades:
                for trade in trades:
                    safe_player = sanitize_csv_value(
                        trade.get("player_name", "Unknown")
                    )
                    trade_type = trade.get("trade_type", "cash")
                    swap_player = trade.get("swap_player", "")
                    swap_player_price = trade.get("swap_player_price", 0)
                    compensation = trade.get("compensation_amount", 0)
                    comp_direction = trade.get("compensation_direction", "")

                    # Format compensation with direction
                    comp_display = ""
                    if compensation and compensation > 0:
                        comp_display = f"{format_amount(compensation)}"
                        if comp_direction:
                            comp_display += f" ({comp_direction})"

                    ts.append(
                        [
                            safe_player,
                            trade.get("from_team", "Unknown"),
                            trade.get("to_team", "Unknown"),
                            format_amount(trade.get("trade_price", 0)),
                            format_amount(trade.get("original_price", 0)),
                            trade_type.upper() if trade_type else "CASH",
                            swap_player if swap_player else "-",
                            (
                                format_amount(swap_player_price)
                                if swap_player_price
                                else "-"
                            ),
                            comp_display if comp_display else "-",
                            trade.get("traded_at", ""),
                        ]
                    )

            # Auto-fit column widths
            ts.column_dimensions["A"].width = 25
            ts.column_dimensions["B"].width = 12
            ts.column_dimensions["C"].width = 12
            ts.column_dimensions["D"].width = 18
            ts.column_dimensions["E"].width = 15
            ts.column_dimensions["F"].width = 12
            ts.column_dimensions["G"].width = 25
            ts.column_dimensions["H"].width = 18
            ts.column_dimensions["I"].width = 20
            ts.column_dimensions["J"].width = 20

            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error updating trade history sheet: {str(e)}")

    @staticmethod
    def update_individual_team_sheets(
        filepath: str,
        teams: Dict[str, int],
        team_squads: Dict[str, List[Tuple[str, int, bool]]],
        retained_players: Dict[str, List[Tuple[str, int]]] = None,
        traded_to_team: Dict[str, Set[str]] = None,
    ) -> None:
        """Update individual team sheets with squad details

        Args:
            filepath: Path to Excel file
            teams: Team purses
            team_squads: All squad data from database (player, price, is_overseas)
            retained_players: Retained players dict (optional, will import if not provided)
            traded_to_team: Dict mapping team code to set of player names traded TO that team
        """
        try:
            if not os.path.exists(filepath):
                FileManager.initialize_excel(filepath)

            # Use a different variable name internally to avoid shadowing
            retained_players_dict = retained_players
            if retained_players_dict is None:
                from retained_players import RETAINED_PLAYERS

                retained_players_dict = RETAINED_PLAYERS

            if traded_to_team is None:
                traded_to_team = {}

            wb = openpyxl.load_workbook(filepath)

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            summary_fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid"
            )
            summary_font = Font(bold=True)

            from config import TEAMS

            sheet_position = 4  # Start after Auction Results, Team Summary, Unsold Players, Trade History

            for team_code in sorted(TEAMS.keys()):

                # Remove existing sheet if present
                if team_code in wb.sheetnames:
                    del wb[team_code]

                # Create new sheet
                ts = wb.create_sheet(team_code, sheet_position)
                sheet_position += 1

                # Headers
                ts.append(["Player Name", "Price", "Type"])
                for cell in ts[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                squad = team_squads.get(team_code, [])
                retained_data = retained_players_dict.get(team_code, [])
                retained_names = {p[0].lower() for p in retained_data}
                traded_names = {p.lower() for p in traded_to_team.get(team_code, set())}

                # Normalize squad to handle both old and new formats
                # New format: (player, price, is_overseas)
                # Old format: (player, price)
                normalized_squad = []
                for item in squad:
                    if len(item) >= 3:
                        normalized_squad.append((item[0], item[1], item[2]))
                    else:
                        normalized_squad.append((item[0], item[1], False))

                # Partition squad
                squad_retained = [
                    (p, pr, is_os)
                    for p, pr, is_os in normalized_squad
                    if p.lower() in retained_names
                ]
                bought_players = [
                    (p, pr, is_os)
                    for p, pr, is_os in normalized_squad
                    if p.lower() not in retained_names and p.lower() not in traded_names
                ]
                traded_players = [
                    (p, pr, is_os)
                    for p, pr, is_os in normalized_squad
                    if p.lower() in traded_names
                ]

                total_spent = sum(pr for _, pr, _ in normalized_squad)
                overseas_count = sum(1 for _, _, is_os in normalized_squad if is_os)

                # Write sections
                row = 2
                if squad_retained:
                    ts.append(["--- Retained ---", "", ""])
                    row += 1
                    for player, price, is_overseas in squad_retained:
                        display_name = f"✈️ {player}" if is_overseas else player
                        ts.append(
                            [
                                sanitize_csv_value(display_name),
                                format_amount(price),
                                "Retained",
                            ]
                        )
                        row += 1

                if bought_players:
                    ts.append(["--- Bought ---", "", ""])
                    row += 1
                    for player, price, is_overseas in bought_players:
                        display_name = f"✈️ {player}" if is_overseas else player
                        ts.append(
                            [
                                sanitize_csv_value(display_name),
                                format_amount(price),
                                "Bought",
                            ]
                        )
                        row += 1

                if traded_players:
                    ts.append(["--- Traded ---", "", ""])
                    row += 1
                    for player, price, is_overseas in traded_players:
                        display_name = f"✈️ {player}" if is_overseas else player
                        ts.append(
                            [
                                sanitize_csv_value(display_name),
                                format_amount(price),
                                "Traded",
                            ]
                        )
                        row += 1

                # Add summary rows
                ts.append([])  # Empty row
                purse_left = teams.get(team_code, 0)

                summary_row = ts.max_row + 1
                ts.append(["Total Players", len(normalized_squad), ""])
                ts.append(["Overseas Players", overseas_count, ""])
                ts.append(["Total Spent", format_amount(total_spent), ""])
                ts.append(["Purse Remaining", format_amount(purse_left), ""])

                # Style summary rows
                for row_num in range(summary_row, ts.max_row + 1):
                    for cell in ts[row_num]:
                        cell.fill = summary_fill
                        cell.font = summary_font

                # Auto-fit column widths
                ts.column_dimensions["A"].width = 30
                ts.column_dimensions["B"].width = 15
                ts.column_dimensions["C"].width = 12

            wb.save(filepath)
        except Exception as e:
            raise Exception(f"Error updating individual team sheets: {str(e)}")

    @staticmethod
    def regenerate_excel_from_db(
        filepath: str,
        sales: List[dict],
        teams: Dict[str, int],
        team_squads: Dict[str, List[Tuple[str, int, bool]]],
        unsold_players: List[Tuple[str, str, Optional[int]]] = None,
        released_players: List[Tuple[str, str, Optional[int]]] = None,
        trades: List[dict] = None,
    ) -> None:
        """Regenerate the entire Excel file from database records.

        Use this after any sale/rollback to ensure Excel matches DB state.

        Args:
            filepath: Path to Excel file
            sales: List of sale records
            teams: Team purses
            team_squads: All squad data (player, price, is_overseas)
            unsold_players: List of (player_name, set_name, base_price) for unsold players
            released_players: List of (player_name, released_from, price) for released players
            trades: List of trade records from trade_history table
        """
        try:
            # Initialize fresh Excel file with all sheets
            FileManager.initialize_excel(filepath)

            wb = openpyxl.load_workbook(filepath)
            sheet = wb["Auction Results"]

            # Add all sales from DB (excluding UNSOLD and RELEASED - they go in separate sheets)
            if sales:
                for sale in sales:
                    team_code = sale.get("team_code", "Unknown")
                    # Skip UNSOLD and RELEASED entries - they go in separate sheets
                    if team_code in ("UNSOLD", "RELEASED"):
                        continue
                    formatted_price = format_amount(sale.get("final_price", 0))
                    timestamp = sale.get(
                        "sold_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                    player_name = sale.get("player_name", "Unknown")

                    # Check if player is overseas from team_squads and add ✈️ emoji
                    is_overseas = False
                    if team_squads and team_code in team_squads:
                        for squad_player in team_squads[team_code]:
                            if squad_player[0].lower() == player_name.lower():
                                is_overseas = (
                                    squad_player[2] if len(squad_player) > 2 else False
                                )
                                break

                    # Add ✈️ emoji for overseas players if not already present
                    display_name = (
                        f"{player_name} ✈️"
                        if is_overseas and "✈️" not in player_name
                        else player_name
                    )

                    # Sanitize player name to prevent formula injection
                    safe_player = sanitize_csv_value(display_name)
                    sheet.append(
                        [
                            safe_player,
                            team_code,
                            formatted_price,
                            timestamp,
                        ]
                    )

            _save_workbook_with_retry(wb, filepath)

            # Update Team Summary
            FileManager.update_team_summary(filepath, teams, team_squads)

            # Update Unsold Players sheet
            if unsold_players:
                FileManager.update_unsold_players_sheet(filepath, unsold_players)

            # Update Released Players sheet
            if released_players:
                FileManager.update_released_players_sheet(filepath, released_players)

            # Update Trade History sheet
            if trades:
                FileManager.update_trade_history_sheet(filepath, trades)

            # Build traded_to_team mapping for individual team sheets
            # This maps team code -> set of player names that were traded TO that team
            traded_to_team: Dict[str, Set[str]] = {}
            if trades:
                for trade in trades:
                    # Main player goes from from_team to to_team
                    to_team = trade.get("to_team")
                    player_name = trade.get("player_name")
                    if to_team and player_name:
                        if to_team not in traded_to_team:
                            traded_to_team[to_team] = set()
                        traded_to_team[to_team].add(player_name)

                    # For swap trades, swap_player goes from to_team to from_team
                    trade_type = trade.get("trade_type", "cash")
                    if trade_type == "swap":
                        swap_player = trade.get("swap_player")
                        from_team = trade.get("from_team")
                        if swap_player and from_team:
                            if from_team not in traded_to_team:
                                traded_to_team[from_team] = set()
                            traded_to_team[from_team].add(swap_player)

            # Update Individual Team Sheets
            FileManager.update_individual_team_sheets(
                filepath, teams, team_squads, traded_to_team=traded_to_team
            )

        except Exception as e:
            raise Exception(f"Error regenerating Excel: {str(e)}")


# -----------------------------------------------------------
# MESSAGE FORMATTER
# -----------------------------------------------------------
class MessageFormatter:

    @staticmethod
    def format_purse_display(teams: Dict[str, int], db=None) -> str:
        """Format team purses with squad info if database is provided"""
        msg = "**Team Summary:**\n```\n"
        msg += f"{'Team':<6} {'Purse':>10} {'Players':>8} {'Overseas':>9}\n"
        msg += "=" * 35 + "\n"

        for t, p in sorted(teams.items()):
            # Get player counts if db is available
            if db:
                try:
                    squad = db.get_team_squad(t)
                    player_count = len(squad)
                    overseas_count = sum(1 for _, _, _, _, is_os in squad if is_os)
                    msg += f"{t:<6} {format_amount(p):>10} {player_count:>8} {overseas_count:>9}\n"
                except Exception:
                    msg += f"{t:<6} {format_amount(p):>10} {'?':>8} {'?':>9}\n"
            else:
                msg += f"{t:<6} {format_amount(p):>10}\n"
        msg += "```"
        return msg

    @staticmethod
    def format_list_display(
        player_lists: Dict[str, List], list_order: List[str]
    ) -> str:
        """Format all lists for display"""
        if not player_lists:
            return "No player lists created yet."

        msg = "**Player Lists:**\n"
        for list_name in list_order:
            if list_name in player_lists:
                players = player_lists[list_name]
                msg += f"\n**{list_name}** ({len(players)} players):\n"
                if players:
                    # Show first 10 players
                    for i, (name, price) in enumerate(players[:10]):
                        price_str = f" - {format_amount(price)}" if price else ""
                        msg += f"  {i+1}. {name}{price_str}\n"
                    if len(players) > 10:
                        msg += f"  ... and {len(players) - 10} more\n"
                else:
                    msg += "  (empty)\n"

        # Show any lists not in order
        for list_name, players in player_lists.items():
            if list_name not in list_order:
                msg += f"\n**{list_name}** ({len(players)} players) [not in order]:\n"
                if players:
                    for i, (name, price) in enumerate(players[:10]):
                        price_str = f" - {format_amount(price)}" if price else ""
                        msg += f"  {i+1}. {name}{price_str}\n"
                    if len(players) > 10:
                        msg += f"  ... and {len(players) - 10} more\n"
                else:
                    msg += "  (empty)\n"

        return msg

    @staticmethod
    def format_player_announcement(
        player: str, base_price: Optional[int], is_overseas: bool = False
    ) -> str:
        player_display = f"{player} ✈️" if is_overseas else player
        return (
            f"**Next Player: {player_display}**\n"
            f"Base Price: {format_amount(base_price)}\n"
            f"Teams can now bid using `/bid`"
        )

    @staticmethod
    def format_bid_message(
        team: str, amount: int, player: str, is_overseas: bool = False
    ) -> str:
        player_display = f"{player} ✈️" if is_overseas else player
        return f"**{team}** bids **{format_amount(amount)}** for **{player_display}**"

    @staticmethod
    def format_sold_message(
        player: str, team: str, amount: int, is_overseas: bool = False
    ) -> str:
        player_display = f"{player} ✈️" if is_overseas else player
        if team == "UNSOLD":
            return (
                f"**UNSOLD**\n"
                f"Player: **{player_display}**\n"
                f"Base Price: **{format_amount(amount)}**"
            )
        return (
            f"**SOLD!**\n"
            f"Player: **{player_display}**\n"
            f"Team: **{team}**\n"
            f"Final Price: **{format_amount(amount)}**"
        )

    @staticmethod
    def format_countdown(seconds: int) -> str:
        return f"**{seconds}** seconds remaining..."

    @staticmethod
    def format_squad_display(
        team_code: str,
        squad: List[Tuple[str, int, str, str, bool]],
        purse: int,
        available_slots: int = 0,
        overseas_slots: int = 0,
    ) -> str:
        """Format a team's squad for display."""
        from config import MAX_SQUAD_SIZE
        from config import MAX_OVERSEAS_LIMIT

        current_players = len(squad)

        # FIX 1: Update sum to unpack 5 values (_, price, _, _, _)
        total_spent = sum(price for _, price, _, _, _ in squad) if squad else 0

        # Calculate overseas count (using index 4)
        overseas_count = sum(1 for row in squad if row[4])

        # Helper function for formatting rows with the airplane symbol
        def fmt_row(p, pr, src, iso):
            symbol = "✈️" if iso else "  "
            base = f"{symbol} {p:25} : {format_amount(pr)}"
            if src:
                base += f" [from {src}]"
            return base

        # FIX 2: Update these comprehensions to unpack 5 items: p, pr, acq, src, iso
        retained_players = [
            (p, pr, src, iso) for p, pr, acq, src, iso in squad if acq == "retained"
        ]
        traded_players = [
            (p, pr, src, iso) for p, pr, acq, src, iso in squad if acq == "traded"
        ]
        bought_players = [
            (p, pr, src, iso) for p, pr, acq, src, iso in squad if acq == "bought"
        ]

        msg = f"**{team_code} Squad:**\n```\n"

        if retained_players:
            msg += "--- Retained ---\n"
            for player, price, source, is_overseas in retained_players:
                msg += f"{fmt_row(player, price, source, is_overseas)}\n"

        if bought_players:
            if retained_players:
                msg += "\n"
            msg += "--- Bought ---\n"
            for player, price, source, is_overseas in bought_players:
                msg += f"{fmt_row(player, price, source, is_overseas)}\n"

        if traded_players:
            if retained_players or bought_players:
                msg += "\n"
            msg += "--- Traded ---\n"
            for player, price, source, is_overseas in traded_players:
                msg += f"{fmt_row(player, price, source, is_overseas)}\n"

        if not retained_players and not bought_players and not traded_players:
            msg += "No players yet.\n"

        # Count bought players (not retained/traded) for slots calculation
        bought_count = len(bought_players)
        slots_remaining = available_slots - bought_count

        # Calculate remaining overseas slots
        remaining_overseas = MAX_OVERSEAS_LIMIT - overseas_count

        msg += f"\n{'='*50}\n"
        msg += f"{'Total Spent':30} : {format_amount(total_spent)}\n"
        msg += f"{'Remaining Purse':30} : {format_amount(purse)}\n"
        msg += f"{'Total Players':30} : {current_players}/{MAX_SQUAD_SIZE}\n"
        msg += f"{'Auction Slots Remaining':30} : {max(0, slots_remaining)}\n"
        msg += f"{'Overseas Players':30} : {overseas_count}/{MAX_OVERSEAS_LIMIT}\n"
        msg += f"{'Overseas Slots Remaining':30} : {max(0, remaining_overseas)}\n"
        msg += "```"
        msg += "\n⚠️ *Manually verify overseas count after trades/releases/buys*"
        return msg


def validate_team_name(team: str, teams: Dict[str, int]) -> Optional[str]:
    team_upper = team.upper()
    return team_upper if team_upper in teams else None


def calculate_next_bid(current_bid: int) -> int:
    from config import get_bid_increment

    return current_bid + get_bid_increment(current_bid)
